# File : results.py

from PySide6.QtWidgets import (QWidget, QLabel, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit,
                               QToolButton,QTableWidget,QDialog, QTableWidgetItem, QHeaderView, QMenu, QFileDialog, QMessageBox)
from PySide6.QtGui import QFont, QIcon, QColor, QBrush
from PySide6.QtCore import Qt, QSize
from ui.widgets.FilterDialog import FilterDialog, apply_filter
from ui.widgets.PlotDialog import PlotDialog
from ui.widgets.FeedbackDialog import FeedbackDialog
from ui.widgets.ConfigurationDialog import ConfigurationDialog
from ui.utils.PathResources import resource_path
from api.feedback_client import FeedbackClient
from api.input_client import InputClient
from PySide6.QtCore import Signal
import fpdf
import json
import datetime
import pickle
from PySide6.QtWidgets import QStyledItemDelegate
from PySide6.QtCore import Qt
import logging
import asyncio
from functools import partial
from PySide6.QtCore import QObject, Signal, Slot
import logging

logger = logging.getLogger(__name__)

class AsyncHelper(QObject):
    """Helper class to run async tasks from Qt"""
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self):
        super().__init__()
        self._loop = None
        self._thread = None

    @Slot(dict)
    def submit_feedback(self, data):
        """Submit feedback and input data asynchronously"""
        async def _submit():
            try:
                feedback_client = FeedbackClient()
                input_client = InputClient()
                
                async with feedback_client, input_client:
                    results = {}
                    
                    if data.get('feedback_data'):
                        feedback_result = await feedback_client.submit_feedback(data['feedback_data'])
                        results['feedback'] = feedback_result
                    
                    if data.get('input_data') and data.get('measurements'):
                        input_result = await input_client.save_input_data({
                            'input_data': data['input_data'],
                            'measurements': data['measurements']
                        })
                        results['input'] = input_result
                    
                    self.finished.emit(results)
            except Exception as e:
                logger.error(f"Error in async operation: {str(e)}")
                self.error.emit(str(e))

        asyncio.create_task(_submit())

class LeftAlignDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = Qt.AlignLeft | Qt.AlignVCenter

class ResultPage(QWidget):
    show_upload_signal = Signal()

    def __init__(self):
        super().__init__()
        self.default_columns = ["Test Name", "Test Number", "Product", "Status",
                              "Mean", "Std"]
        self.visible_columns = self.default_columns + ["ACTION"]
        self.async_helper = AsyncHelper()
        self.async_helper.finished.connect(self._on_submit_complete)
        self.async_helper.error.connect(self._on_submit_error)
        self.all_columns = []
        self.initUI()

    def initUI(self):
        self.setFixedWidth(1200)
        self.setStyleSheet("background-color: white; color: black")
        mainLayout = QVBoxLayout(self)
        mainLayout.setSpacing(0)
        row1 = QHBoxLayout()
        titleLayout = QVBoxLayout()
        title = QLabel("Dashboard")
        title.setFont(QFont("Arial", 12, QFont.Bold))
        title.setStyleSheet("margin-bottom: 0px;")
        subtitle = QLabel("Your files are being processed")
        subtitle.setFont(QFont("Arial", 8))
        subtitle.setStyleSheet("color: gray; margin-left: 0px; margin-top: 5px; margin-bottom: 20px;")

        titleLayout.addWidget(title)
        titleLayout.addWidget(subtitle)
        
        row1.addLayout(titleLayout)
        row1.addStretch()

        self.restartButton = QPushButton("Restart") 
        self.restartButton.setStyleSheet("""
            QPushButton {
                background-color: #1849D6;
                color: white;
                border-radius: 5px;
                padding: 0px 15px;
                max-height : 27px;
            }
        """)
        self.restartButton.clicked.connect(self.restartProcessing)
        
        self.exportButton = QToolButton()
        self.exportButton.setText("Export")
        self.exportButton.setStyleSheet("""
            QToolButton {
                background-color: #1FBE42;
                color: white;
                border-radius: 5px;
                padding: 0px 15px;
                max-height : 27px;
                margin-right:5px;
            }
        """)
        
        self.exportMenu = QMenu()
        self.exportMenu.addAction("CSV", self.exportCSV)
        self.exportMenu.addAction("Excel", self.exportExcel)
        self.exportMenu.addAction("PDF", self.exportPDF)
        self.exportMenu.addAction("HTML", self.exportHTML)
        self.exportButton.setMenu(self.exportMenu)
        self.exportButton.setPopupMode(QToolButton.InstantPopup)

        self.exportDashboardButton = QPushButton("Export Dashboard")
        self.exportDashboardButton.setStyleSheet("""
            QPushButton {
                background-color: #17A2B8;
                color: white;
                border-radius: 5px;
                padding: 0px 15px;
                max-height: 27px;
                margin-right: 5px;
            }
        """)
        self.exportDashboardButton.clicked.connect(self.exportDashboard)

        self.configButton = QPushButton("Configuration")
        self.configButton.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border-radius: 5px;
                padding: 0px 15px;
                max-height: 27px;
                margin-right: 5px;
            }
        """)
        self.configButton.clicked.connect(self.showConfiguration)
        
        row1.addWidget(self.exportButton, alignment=Qt.AlignRight)
        row1.addWidget(self.exportDashboardButton, alignment=Qt.AlignRight)
        row1.addWidget(self.configButton, alignment=Qt.AlignRight)
        row1.addWidget(self.restartButton, alignment=Qt.AlignRight)

        searchLayout = QHBoxLayout()
        searchBoxLayout = QHBoxLayout()
        
        searchContainer = QWidget()
        searchContainerLayout = QHBoxLayout(searchContainer)
        searchContainerLayout.setContentsMargins(15,3,3,3)
        searchContainer.setStyleSheet("""
                background-color: rgba(24, 73, 214, 0.08);
                border-radius: 19px;
        """)
        self.searchBox = QLineEdit()
        self.searchBox.setPlaceholderText("Search...")
        self.searchBox.setStyleSheet("""
            QLineEdit {
                margin: 0;
                background-color: transparent;
            }
        """)
        self.searchBox.setFixedHeight(30)
        self.searchBox.textChanged.connect(self.searchTable)
        
        searchIcon = QToolButton()
        searchIcon.setIcon(QIcon(resource_path("./resources/icons/search.png")))
        searchIcon.setIconSize(QSize(30, 30))
        searchIcon.setStyleSheet("background: transparent; border: none; margin: 0")
        
        searchContainerLayout.addWidget(self.searchBox)
        searchContainerLayout.addWidget(searchIcon)
        
        filterIcon = QToolButton()
        filterIcon.setIcon(QIcon(resource_path("./resources/icons/filter.png")))
        filterIcon.setIconSize(QSize(30, 30))
        filterIcon.setStyleSheet("background: transparent; border: none; margin: 0")
        filterIcon.clicked.connect(self.filterTable)
        
        searchBoxLayout.addWidget(searchContainer)
        searchLayout.addLayout(searchBoxLayout)
        searchLayout.addWidget(filterIcon)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "Test Name", "Test Number", "Product", "Status",
                              "Mean", "Std", "ACTION"
        ])
        
        self.table.setShowGrid(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                gridline-color: #E5E7EB;
                background-color: white;
            }
            QTableWidget::item {
                border-right: 1px solid #E5E7EB;
                padding: 5px 8px;
            }
            QTableWidget QTableCornerButton::section {
                background-color: white;
                border: none;
            }
        """)
        
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: white;
                padding: 5px 8px;
                border: none;
                border-right: 1px solid #E5E7EB;
                border-bottom: 1px solid #E5E7EB;
                font-weight: bold;
                text-align: left;
            }
            QHeaderView::section:first {
                border-left: none;
            }
        """)
        
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        
        self.table.setItemDelegate(LeftAlignDelegate())
        
        row1.setContentsMargins(0,0,0,20)
        mainLayout.addLayout(row1)
        searchLayout.setContentsMargins(0,0,0,20)
        mainLayout.addLayout(searchLayout)
        mainLayout.addWidget(self.table)
        mainLayout.setContentsMargins(40,40,40,40)

    def apply_status_styling(self, item, status):
        """Apply color styling to status cell"""
        item = QTableWidgetItem(status)
        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        status = status.strip().lower()
        if status == "similar distribution":
            item.setForeground(QBrush(QColor("#1E7D34")))
            item.setFont(QFont("Arial", weight=QFont.Bold))
        elif status == "moderately similar":
            item.setForeground(QBrush(QColor("#B88A00")))
            item.setFont(QFont("Arial", weight=QFont.Bold))
        elif status == "completely different":
            item.setForeground(QBrush(QColor("#D9534F")))
            item.setFont(QFont("Arial", weight=QFont.Bold))
        
        return item

    def showConfiguration(self):
        config_dialog = ConfigurationDialog(self.all_columns, self.visible_columns, self)
        if config_dialog.exec() == QDialog.Accepted:
            self.visible_columns = config_dialog.get_selected_columns()
            self.updateTableColumns()
            if hasattr(self, 'data'):
                self.populateTable(self.data)

    def exportDashboard(self):
        """Export complete dashboard state as pickle file"""
        path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export Dashboard", 
            f"dashboard_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pkl", 
            "Dashboard Files (*.pkl)"
        )
        if path:
            try:
                dashboard_data = {
                    'data': self.data if hasattr(self, 'data') else None,
                    'visible_columns': self.visible_columns,
                    'all_columns': self.all_columns,
                    'table_data': self._extract_table_data(),
                    'metadata': {
                        'export_date': datetime.datetime.now().isoformat(),
                        'version': '1.0',
                        'rows': self.table.rowCount(),
                        'columns': self.table.columnCount()
                    }
                }
                
                with open(path, 'wb') as f:
                    pickle.dump(dashboard_data, f)
                
                QMessageBox.information(
                    self,
                    "Export Successful",
                    f"Dashboard exported successfully to:\n{path}"
                )
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Export Failed",
                    f"Failed to export dashboard:\n{str(e)}"
                )

    def _extract_table_data(self):
        """Extract all table data for export"""
        table_data = []
        for row in range(self.table.rowCount()):
            row_data = {}
            for col in range(self.table.columnCount() - 1):
                header = self.table.horizontalHeaderItem(col)
                if header:
                    item = self.table.item(row, col)
                    row_data[header.text()] = item.text() if item else ""
            table_data.append(row_data)
        return table_data

    def populateTable(self, df):
        if df is None:
            return
        
        self.data = df
        
        self.all_columns = [col for col in df.columns if not (
            col.endswith('_reference') or 
            col in ['input_data', 'reference_data'] or
            '_reference' in col
        )]
        
        for col in self.default_columns:
            if col not in self.visible_columns:
                self.visible_columns.append(col)
        
        if "ACTION" in self.visible_columns:
            self.visible_columns.remove("ACTION")
        self.visible_columns.append("ACTION")
        
        self.table.setRowCount(len(self.data))
        
        for i, (index, row) in enumerate(self.data.iterrows()):
            for col_index, column in enumerate(self.visible_columns):
                if column == "ACTION":
                    action_widget = self.create_action_buttons(i)
                    self.table.setCellWidget(i, col_index, action_widget)
                else:
                    value = str(row.get(column, ""))
                    item = QTableWidgetItem(value)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    
                    if column == "Status":
                        item = self.apply_status_styling(item, value)
                    
                    self.table.setItem(i, col_index, item)

    def updateTableColumns(self):
        if "ACTION" in self.visible_columns:
            self.visible_columns.remove("ACTION")
        self.visible_columns.append("ACTION")
        
        self.table.clear()
        self.table.setColumnCount(len(self.visible_columns))
        self.table.setHorizontalHeaderLabels(self.visible_columns)
        
        for i, column in enumerate(self.visible_columns):
            if column == "ACTION":
                self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Fixed)
                self.table.setColumnWidth(i, 100)
            else:
                self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
        
        if hasattr(self, 'data'):
            self.populateTable(self.data)

    def create_action_buttons(self, row_idx):
        """Create action buttons for a row"""
        action_layout = QHBoxLayout()
        action_layout.setContentsMargins(2, 2, 2, 2)
        
        viz_button = QToolButton()
        viz_button.setIcon(QIcon(resource_path("./resources/icons/Plot.png")))
        viz_button.setIconSize(QSize(40, 40))
        viz_button.setToolTip("Visualize")
        viz_button.clicked.connect(lambda checked, r=row_idx: self.showPlot(r))
        
        edit_button = QToolButton()
        edit_button.setIcon(QIcon(resource_path("./resources/icons/edit.png")))
        edit_button.setIconSize(QSize(40, 40))
        edit_button.clicked.connect(lambda _, r=row_idx: self.editRow(r))
        
        action_layout.addWidget(viz_button)
        action_layout.addWidget(edit_button)
        
        action_widget = QWidget()
        action_widget.setLayout(action_layout)
        
        return action_widget

    def showPlot(self, row_idx):
        """Show distribution plot for the selected row"""
        row_data = self.data.iloc[row_idx]
        input_data = row_data['input_data']
        reference_data = row_data['reference_data']
        test_name = row_data['Test Name']
        if input_data is not None and reference_data is not None:
            plot_dialog = PlotDialog(input_data, reference_data, test_name, self)
            plot_dialog.exec_()

    def editRow(self, row):
        try:
            df_row = self.data.iloc[row]
            
            status_col_idx = self.visible_columns.index("Status")
            test_name_col_idx = self.visible_columns.index("Test Name")
            current_status = self.table.item(row, status_col_idx).text()
            test_name = self.table.item(row, test_name_col_idx).text()
            test_number = df_row.get('Test Number', '')
            
            input_data = {
                'input_data': df_row.get('input_data'),
                'lsl': df_row.get('LSL'),
                'usl': df_row.get('USL'),
                'measurements': df_row.get('measurements', [])
            }
            
            dialog = FeedbackDialog(
                current_status=current_status,
                test_name=test_name,
                test_number=str(test_number),
                lot=df_row.get('lot_input', ''),
                insertion=df_row.get('insertion_input', ''),
                initial_label=current_status,
                reference_id=df_row.get('reference_id', ''),
                input_id=df_row.get('input_id', ''),
                input_data=input_data,
                parent=self
            )
            
            if dialog.exec() == QDialog.Accepted:
                values = dialog.get_values()
                new_status = values['new_status']
                
                status_item = self.apply_status_styling(None, new_status)
                self.table.setItem(row, status_col_idx, status_item)
                
                if values['send_feedback'] and values['feedback_data']:
                    feedback_client = FeedbackClient()
                    input_client = InputClient()
                    
                    try:
                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        
                        loop.run_until_complete(feedback_client.submit_feedback(values['feedback_data']))
                        
                        if values['input_data'] and values['measurements']:
                            input_save_data = {
                                **values['input_data'],
                                'measurements': values['measurements']
                            }
                            loop.run_until_complete(input_client.save_input_data(input_save_data))
                        
                        loop.close()
                        
                        QMessageBox.information(
                            self,
                            "Success",
                            "Status updated and all data submitted successfully!"
                        )
                        
                    except Exception as e:
                        logger.error(f"Error submitting data: {str(e)}")
                        QMessageBox.warning(
                            self,
                            "Warning",
                            f"Status updated but data submission failed: {str(e)}"
                        )
                else:
                    QMessageBox.information(
                        self,
                        "Success",
                        "Status updated successfully!"
                    )
                        
        except Exception as e:
            logger.error(f"Error editing row: {str(e)}")
            QMessageBox.critical(
                self,
                "Error",
                f"Error editing row: {str(e)}"
            )

    @Slot(dict)
    def _on_submit_complete(self, results):
        """Handle completion of async submission"""
        QMessageBox.information(
            self,
            "Success",
            "Status updated and all data submitted successfully!"
        )

    @Slot(str)
    def _on_submit_error(self, error_msg):
        """Handle error in async submission"""
        QMessageBox.warning(
            self,
            "Warning",
            f"Status updated but data submission failed: {error_msg}"
        )

    def removeRow(self, row):
        """Handle removal of a row"""
        try:
            self.table.removeRow(row)
            self.data = self.data.drop(self.data.index[row])
        except Exception as e:
            print(f"Error removing row: {str(e)}")

    def exportCSV(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV", "", "CSV Files (*.csv)")
        if path:
            with open(path, 'w') as file:
                headers = []
                for column in range(self.table.columnCount() - 1):
                    header = self.table.horizontalHeaderItem(column)
                    if header:
                        headers.append(header.text())
                file.write(",".join(headers) + "\n")
                
                for row in range(self.table.rowCount()):
                    rowData = []
                    for column in range(self.table.columnCount() - 1):
                        item = self.table.item(row, column)
                        if item:
                            rowData.append(item.text())
                    file.write(",".join(rowData) + "\n")

    def exportExcel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if path:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                
                wb = Workbook()
                ws = wb.active
                
                headers = []
                bold_font = Font(bold=True)
                for column in range(self.table.columnCount() - 1):
                    header = self.table.horizontalHeaderItem(column)
                    if header:
                        cell = ws.cell(row=1, column=column + 1, value=header.text())
                        cell.font = bold_font
                
                for row in range(self.table.rowCount()):
                    row_data = []
                    for column in range(self.table.columnCount() - 1):
                        item = self.table.item(row, column)
                        if item:
                            row_data.append(item.text())
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row + 2, column=col, value=value)
                
                wb.save(path)
            except ImportError:
                QMessageBox.warning(self, "Export Failed", "Openpyxl module not found. Please install it to export to Excel.")

    def exportPDF(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", "", "PDF Files (*.pdf)")
        if path:
            try:
                from fpdf import FPDF
                
                class PDF(FPDF):
                    def header(self):
                        self.set_font('Arial', 'B', 15)
                        self.cell(80)
                        self.cell(30, 10, 'Test Results', 0, 0, 'C')
                        self.ln(20)

                pdf = PDF()
                pdf.add_page()
                
                pdf.set_font("Arial", 'B', 10)
                
                col_width = pdf.w / (self.table.columnCount() - 1)
                
                for column in range(self.table.columnCount() - 1):
                    header = self.table.horizontalHeaderItem(column)
                    if header:
                        pdf.cell(col_width, 10, str(header.text()), 1)
                pdf.ln()
                
                pdf.set_font("Arial", size=10)
                
                for row in range(self.table.rowCount()):
                    for column in range(self.table.columnCount() - 1):
                        item = self.table.item(row, column)
                        if item:
                            pdf.cell(col_width, 10, str(item.text()), 1)
                    pdf.ln()
                
                pdf.output(path)
            except ImportError:
                QMessageBox.warning(self, "Export Failed", "FPDF module not found. Please install it to export to PDF.")

    def exportHTML(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save HTML", "", "HTML Files (*.html)")
        if path:
            with open(path, 'w') as file:
                file.write("""
                <html>
                <head>
                    <style>
                        table { border-collapse: collapse; width: 100%; }
                        th, td { border: 1px solid black; padding: 8px; text-align: left; }
                        th { background-color: #f2f2f2; }
                    </style>
                </head>
                <body>
                <table>
                <thead>
                    <tr>
                """)
                
                for column in range(self.table.columnCount() - 1):
                    header = self.table.horizontalHeaderItem(column)
                    if header:
                        file.write(f"<th>{header.text()}</th>\n")
                file.write("</tr></thead>\n<tbody>\n")
                
                for row in range(self.table.rowCount()):
                    file.write("<tr>\n")
                    for column in range(self.table.columnCount() - 1):
                        item = self.table.item(row, column)
                        if item:
                            file.write(f"<td>{item.text()}</td>\n")
                    file.write("</tr>\n")
                
                file.write("</tbody></table>\n</body></html>")

    def searchTable(self):
        query = self.searchBox.text().lower()
        for row in range(self.table.rowCount()):
            match = False
            for column in range(self.table.columnCount() - 1):
                item = self.table.item(row, column)
                if item and query in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match)

    def filterTable(self):
        """Handle filtering of the table"""
        filterDialog = FilterDialog(self)
        if filterDialog.exec() == QDialog.Accepted:
            filter_conditions = filterDialog.get_filter_values()
            apply_filter(self.table, filter_conditions)

    def restartProcessing(self):
        """Handle restart button click"""
        reply = QMessageBox.question(self, "Restart", 
                                   "Are you sure you want to restart? This will take you back to the upload page.",
                                   QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.show_upload_signal.emit()